"""Export orders and metrics to an Excel workbook with live formulas."""

from __future__ import annotations

from collections import defaultdict
from datetime import date
from pathlib import Path
from typing import List

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows

from .metrics import total_revenue, total_units
from .schema import UnifiedOrder


def export_excel(orders: List[UnifiedOrder], path: Path) -> Path:
    wb = Workbook()

    # Orders sheet
    ws_orders = wb.active
    ws_orders.title = "Orders"
    order_rows = []
    for o in orders:
        order_rows.append(
            {
                "platform": o.platform,
                "order_id": o.order_id,
                "date": o.order_date,
                "product": o.product_name,
                "sku": o.sku,
                "quantity": o.quantity,
                "unit_price": o.unit_price,
                "currency": o.currency,
                "revenue": o.revenue,
                "fees": o.fees,
                "shipping": o.shipping_cost,
                "refund": o.refund,
                "country": o.country or "",
                "base_currency": o.base_currency,
                "revenue_base": o.revenue_base if o.revenue_base is not None else o.revenue,
            }
        )
    df_orders = pd.DataFrame(order_rows)
    for r_idx, row in enumerate(dataframe_to_rows(df_orders, index=False, header=True), 1):
        ws_orders.append(row)
    ws_orders.freeze_panes = "A2"

    # Summary sheet
    ws_summary = wb.create_sheet("Summary")
    ws_summary.append(["Metric", "Value"])
    ws_summary.append(["Total revenue (base)", f"=SUM(Orders!O2:O{len(order_rows) + 1})"])
    ws_summary.append(["Total units", f"=SUM(Orders!F2:F{len(order_rows) + 1})"])
    ws_summary.append(["Refund count", f"=COUNTIF(Orders!K2:K{len(order_rows) + 1},TRUE)"])
    ws_summary["A1"].font = Font(bold=True)
    ws_summary["B1"].font = Font(bold=True)

    # Products sheet with formulas referencing Orders
    ws_products = wb.create_sheet("Products")
    product_revenue: defaultdict[str, float] = defaultdict(float)
    product_units: defaultdict[str, int] = defaultdict(int)
    for o in orders:
        if o.refund:
            continue
        product_revenue[o.product_name] += o.revenue_base if o.revenue_base is not None else o.revenue
        product_units[o.product_name] += o.quantity

    ws_products.append(["Product", "Units", "Revenue (base)"])
    for name in sorted(product_revenue):
        ws_products.append(
            [
                name,
                f'=SUMIF(Orders!D2:D{len(order_rows) + 1},"{name}",Orders!F2:F{len(order_rows) + 1})',
                f'=SUMIF(Orders!D2:D{len(order_rows) + 1},"{name}",Orders!O2:O{len(order_rows) + 1})',
            ]
        )
    ws_products["A1"].font = Font(bold=True)
    ws_products["B1"].font = Font(bold=True)
    ws_products["C1"].font = Font(bold=True)

    wb.save(str(path))
    return path
